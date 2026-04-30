#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
双扣 - 八王千变 发牌器 - Excel 统计导出
支持批量模拟和统计分析
"""

import json
import random
from datetime import datetime
from pathlib import Path
from typing import List, Dict
from collections import Counter, defaultdict

# 尝试导入 openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    print("⚠️ 未安装 openpyxl，正在安装...")
    import subprocess
    subprocess.check_call(['pip3', 'install', 'openpyxl', '-q'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True

from dealer import ShuangkouDealer, CARD_RANKS


# ==================== 样式定义 ====================

class StyleConfig:
    """Excel 样式配置"""
    
    # 字体
    TITLE_FONT = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
    HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    NORMAL_FONT = Font(name='微软雅黑', size=10)
    
    # 对齐
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
    
    # 填充色
    TITLE_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    HEADER_FILL = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    ALT_FILL = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
    
    # 边框
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


# ==================== Excel 导出器 ====================

class ExcelExporter:
    """Excel 统计报告导出器"""
    
    def __init__(self, config: Dict):
        self.config = config
        self.rule_name = config.get('rule_name', 'unknown')
        self.rule_abbr = config.get('rule_abbr', 'unknown')
        self.wb = Workbook()
        # 保留默认 sheet，重命名为统计总览
        ws = self.wb.active
        ws.title = '📊 统计总览'
        
    def generate_filename(self) -> str:
        """生成文件名（包含规则缩写和时间戳）"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"双扣统计_{self.rule_abbr}_{timestamp}.xlsx"
    
    def create_summary_sheet(self, results: List[Dict], total_games: int):
        """创建汇总分析表"""
        ws = self.wb.create_sheet('📊 汇总分析')
        
        # 标题
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f'双扣 - 八王千变 发牌统计报告\n规则：{self.rule_name}'
        title_cell.font = StyleConfig.TITLE_FONT
        title_cell.fill = StyleConfig.TITLE_FILL
        title_cell.alignment = StyleConfig.CENTER_ALIGN
        
        # 基本信息
        ws.merge_cells('A2:H2')
        info_cell = ws['A2']
        info_cell.value = f'总对局数：{total_games} | 生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        info_cell.font = StyleConfig.NORMAL_FONT
        info_cell.alignment = StyleConfig.CENTER_ALIGN
        
        # 统计指标
        headers = ['统计指标', '玩家 1', '玩家 2', '玩家 3', '玩家 4', '总计', '平均', '最大']
        ws.append(headers)
        
        # 设置表头样式
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col)
            cell.font = StyleConfig.HEADER_FONT
            cell.fill = StyleConfig.HEADER_FILL
            cell.alignment = StyleConfig.CENTER_ALIGN
            cell.border = StyleConfig.THIN_BORDER
        
        # 计算统计数据
        stats = {
            '炸弹总数': [],
            '万能牌数': [],
            '顺子数量': [],
            '对子数量': [],
            '三条数量': [],
            '贡献分': [],
            '连炸贡献分': [],
            '单炸贡献分': [],
        }
        for line in range(4, 17):
            stats[f'{line}线'] = []
        
        for result in results:
            for i in range(4):
                bombs = result['players'][i]['bombs']
                jokers = result['players'][i]['jokers']
                sequences = result['players'][i].get('sequences', 0)
                pairs = result['players'][i].get('pairs', 0)
                triplets = result['players'][i].get('triplets', 0)
                score = result['players'][i].get('contribution_score', 0)
                
                stats['炸弹总数'].append(len(bombs))
                stats['万能牌数'].append(jokers)
                # 计算线数分布
                line_counts = {}
                for line in range(4, 17):
                    line_counts[line] = 0
                
                # 区分王炸和普通炸弹
                joker_bombs = [b for b in bombs if b['rank'] in ['👑', '🃏']]
                normal_bombs = [b for b in bombs if b['rank'] not in ['👑', '🃏']]
                
                # 王炸线数 = 王数 + 3
                for jb in joker_bombs:
                    line = jb['size'] + 3
                    if 4 <= line <= 16:
                        line_counts[line] += 1
                
                # 普通炸弹检测连炸
                if normal_bombs:
                    CARD_RANKS = ['3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K', 'A', '2']
                    rank_indices = sorted([CARD_RANKS.index(b['rank']) for b in normal_bombs if b['rank'] in CARD_RANKS])
                    groups = []
                    if rank_indices:
                        current = [rank_indices[0]]
                        for j in range(1, len(rank_indices)):
                            if rank_indices[j] == rank_indices[j-1] + 1:
                                current.append(rank_indices[j])
                            else:
                                groups.append(current)
                                current = [rank_indices[j]]
                        groups.append(current)
                    
                    for group in groups:
                        if len(group) >= 3:
                            # 连炸线数 = 最小炸弹张数 + 连环数
                            min_size = min(b['size'] for b in normal_bombs if CARD_RANKS.index(b['rank']) in group)
                            line = min_size + len(group)
                            if 4 <= line <= 16:
                                line_counts[line] += 1
                        else:
                            # 单个炸弹
                            for idx in group:
                                b_size = next(b['size'] for b in normal_bombs if CARD_RANKS.index(b['rank']) == idx)
                                if 4 <= b_size <= 16:
                                    line_counts[b_size] += 1
                
                for line in range(4, 17):
                    stats[f'{line}线'].append(line_counts[line])
                stats['顺子数量'].append(sequences)
                stats['对子数量'].append(pairs)
                stats['三条数量'].append(triplets)
                stats['贡献分'].append(score)
                detail = result['players'][i].get('contribution_detail', {})
                stats['连炸贡献分'].append(detail.get('chain_score', 0))
                stats['单炸贡献分'].append(detail.get('single_score', 0))
        
        # 写入统计数据
        row = 4
        for stat_name, values in stats.items():
            player_values = [values[i*4:(i+1)*4] for i in range(total_games)]
            
            # 每个玩家的平均值
            for i in range(4):
                player_vals = [pv[i] for pv in player_values]
                ws.cell(row=row, column=i+2).value = sum(player_vals) / total_games
            
            # 总计和平均
            all_values = [v for sublist in player_values for v in sublist]
            ws.cell(row=row, column=6).value = sum(all_values)
            ws.cell(row=row, column=7).value = sum(all_values) / (total_games * 4)
            ws.cell(row=row, column=8).value = max(all_values)
            
            # 指标名称
            ws.cell(row=row, column=1).value = stat_name
            
            # 设置样式
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.font = StyleConfig.NORMAL_FONT
                cell.alignment = StyleConfig.CENTER_ALIGN
                cell.border = StyleConfig.THIN_BORDER
                if row % 2 == 0:
                    cell.fill = StyleConfig.ALT_FILL
            
            row += 1
        
        # 调整列宽
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 12
        ws.column_dimensions['A'].width = 15
    
    def create_distribution_sheet(self, results: List[Dict]):
        """创建炸弹大小分布表"""
        ws = self.wb.create_sheet('📈 炸弹分布')
        
        # 标题
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = '炸弹大小分布统计'
        title_cell.font = StyleConfig.TITLE_FONT
        title_cell.fill = StyleConfig.TITLE_FILL
        title_cell.alignment = StyleConfig.CENTER_ALIGN
        
        # 表头
        headers = ['炸弹大小', '出现次数', '占比', '平均每局', '说明']
        ws.append(headers)
        
        # 设置表头样式
        for col in range(1, 6):
            cell = ws.cell(row=2, column=col)
            cell.font = StyleConfig.HEADER_FONT
            cell.fill = StyleConfig.HEADER_FILL
            cell.alignment = StyleConfig.CENTER_ALIGN
            cell.border = StyleConfig.THIN_BORDER
        
        # 统计炸弹线数分布（考虑连炸）
        CARD_RANKS = ['3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K', 'A', '2']
        line_counter = {}
        for line in range(4, 17):
            line_counter[line] = 0
        
        for result in results:
            for player in result['players']:
                bombs = player['bombs']
                joker_bombs = [b for b in bombs if b['rank'] in ['👑', '🃏']]
                normal_bombs = [b for b in bombs if b['rank'] not in ['👑', '🃏']]
                
                for jb in joker_bombs:
                    line = jb['size'] + 3
                    if 4 <= line <= 16:
                        line_counter[line] += 1
                
                if normal_bombs:
                    rank_indices = sorted([CARD_RANKS.index(b['rank']) for b in normal_bombs if b['rank'] in CARD_RANKS])
                    groups = []
                    if rank_indices:
                        current = [rank_indices[0]]
                        for j in range(1, len(rank_indices)):
                            if rank_indices[j] == rank_indices[j-1] + 1:
                                current.append(rank_indices[j])
                            else:
                                groups.append(current)
                                current = [rank_indices[j]]
                        groups.append(current)
                    
                    for group in groups:
                        if len(group) >= 3:
                            min_size = min(b['size'] for b in normal_bombs if CARD_RANKS.index(b['rank']) in group)
                            line = min_size + len(group)
                            if 4 <= line <= 16:
                                line_counter[line] += 1
                        else:
                            for idx in group:
                                b_size = next(b['size'] for b in normal_bombs if CARD_RANKS.index(b['rank']) == idx)
                                if 4 <= b_size <= 16:
                                    line_counter[b_size] += 1
        
        total_bombs = sum(line_counter.values())
        desc_map = {4: '普通炸弹 (无贡献分)', 5: '5线 (×2)', 6: '6线 (×4)'}
        for line in range(7, 17):
            desc_map[line] = f'{line}线 (×{2**(line-4)})'
        
        row = 3
        for line in range(4, 17):
            count = line_counter[line]
            if count == 0:
                continue
            percentage = count / total_bombs * 100 if total_bombs > 0 else 0
            per_game = count / len(results)
            desc = desc_map.get(line, '')
            
            ws.append([f'{line}线', count, f'{percentage:.2f}%', f'{per_game:.2f}', desc])
            
            # 设置样式
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.font = StyleConfig.NORMAL_FONT
                cell.alignment = StyleConfig.CENTER_ALIGN
                cell.border = StyleConfig.THIN_BORDER
                if row % 2 == 1:
                    cell.fill = StyleConfig.ALT_FILL
            
            row += 1
        
        # 调整列宽
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def create_team_balance_sheet(self, results: List[Dict]):
        """创建队伍平衡分析表"""
        ws = self.wb.create_sheet('⚖️ 队伍平衡')
        
        # 标题
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = '队伍牌力平衡分析'
        title_cell.font = StyleConfig.TITLE_FONT
        title_cell.fill = StyleConfig.TITLE_FILL
        title_cell.alignment = StyleConfig.CENTER_ALIGN
        
        # 表头
        headers = ['队伍', '平均炸弹数', '平均万能牌数', '胜率估算']
        ws.append(headers)
        
        # 设置表头样式
        for col in range(1, 5):
            cell = ws.cell(row=2, column=col)
            cell.font = StyleConfig.HEADER_FONT
            cell.fill = StyleConfig.HEADER_FILL
            cell.alignment = StyleConfig.CENTER_ALIGN
            cell.border = StyleConfig.THIN_BORDER
        
        # 统计队伍数据
        team0_bombs = []
        team1_bombs = []
        team0_jokers = []
        team1_jokers = []
        
        for result in results:
            team0_bomb = sum(len(result['players'][i]['bombs']) for i in [0, 2])
            team1_bomb = sum(len(result['players'][i]['bombs']) for i in [1, 3])
            team0_joker = sum(result['players'][i]['jokers'] for i in [0, 2])
            team1_joker = sum(result['players'][i]['jokers'] for i in [1, 3])
            
            team0_bombs.append(team0_bomb)
            team1_bombs.append(team1_bomb)
            team0_jokers.append(team0_joker)
            team1_jokers.append(team1_joker)
        
        # 写入数据
        row = 3
        for team_name, bombs, jokers in [
            ('队伍 0 (玩家 1、3)', team0_bombs, team0_jokers),
            ('队伍 1 (玩家 2、4)', team1_bombs, team1_jokers)
        ]:
            avg_bombs = sum(bombs) / len(results)
            avg_jokers = sum(jokers) / len(results)
            
            # 简单的胜率估算（基于炸弹数）
            total_bombs = sum(team0_bombs) + sum(team1_bombs)
            if total_bombs > 0:
                win_rate = sum(bombs) / total_bombs * 100
            else:
                win_rate = 50
            
            ws.append([team_name, f'{avg_bombs:.2f}', f'{avg_jokers:.2f}', f'{win_rate:.1f}%'])
            
            # 设置样式
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = StyleConfig.NORMAL_FONT
                cell.alignment = StyleConfig.CENTER_ALIGN
                cell.border = StyleConfig.THIN_BORDER
                if row % 2 == 1:
                    cell.fill = StyleConfig.ALT_FILL
            
            row += 1
        
        # 调整列宽
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def create_sample_sheet(self, results: List[Dict], sample_size: int = 5):
        """创建典型牌例展示表"""
        ws = self.wb.create_sheet('🎴 典型牌例')
        
        # 标题
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = f'典型牌例展示（随机抽取 {sample_size} 局）'
        title_cell.font = StyleConfig.TITLE_FONT
        title_cell.fill = StyleConfig.TITLE_FILL
        title_cell.alignment = StyleConfig.CENTER_ALIGN
        
        # 随机抽取 sample_size 局
        samples = random.sample(results, min(sample_size, len(results)))
        
        for idx, result in enumerate(samples):
            # 局数标题
            row = idx * 10 + 2
            ws.merge_cells(f'A{row}:C{row}')
            game_title = ws[f'A{row}']
            game_title.value = f'第 {idx + 1} 局'
            game_title.font = Font(name='微软雅黑', size=12, bold=True)
            game_title.alignment = StyleConfig.LEFT_ALIGN
            
            # 每个玩家的牌
            for i, player in enumerate(result['players']):
                row = idx * 10 + 3 + i * 2
                player_name = f'玩家{i+1} (队伍{player["team"]})'
                bombs_str = ', '.join([f"{b['rank'] * b['size']}" for b in player['bombs']])
                jokers_str = f"万能牌×{player['jokers']}"
                
                ws.cell(row=row, column=1).value = player_name
                ws.cell(row=row, column=2).value = f'炸弹：{len(player["bombs"])}个'
                ws.cell(row=row, column=3).value = jokers_str
                
                row = idx * 10 + 4 + i * 2
                ws.cell(row=row, column=1).value = f'炸弹详情：{bombs_str if bombs_str else "无"}'
                ws.merge_cells(f'A{row}:C{row}')
        
        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
    
    def export(self, results: List[Dict], output_dir: str = '.') -> str:
        """导出 Excel 文件（所有统计整合到一个 sheet）"""
        ws = self.wb.active  # 使用默认的统计总览 sheet
        
        # 标题
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f'双扣 - 八王千变 发牌统计报告\n规则：{self.rule_name}'
        title_cell.font = StyleConfig.TITLE_FONT
        title_cell.fill = StyleConfig.TITLE_FILL
        title_cell.alignment = StyleConfig.CENTER_ALIGN
        
        # 基本信息
        ws.merge_cells('A2:H2')
        info_cell = ws['A2']
        info_cell.value = f'总对局数：{len(results)} | 生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        info_cell.font = StyleConfig.NORMAL_FONT
        info_cell.alignment = StyleConfig.CENTER_ALIGN
        
        # === 第一部分：汇总分析 ===
        row = 4
        ws.merge_cells(f'A{row}:H{row}')
        section_title = ws[f'A{row}']
        section_title.value = '📊 汇总分析'
        section_title.font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')
        section_title.alignment = StyleConfig.LEFT_ALIGN
        
        # 统计指标表头
        row += 1
        headers = ['统计指标', '玩家 1', '玩家 2', '玩家 3', '玩家 4', '总计', '平均', '最大']
        ws.append(headers)
        
        # 设置表头样式
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.font = StyleConfig.HEADER_FONT
            cell.fill = StyleConfig.HEADER_FILL
            cell.alignment = StyleConfig.CENTER_ALIGN
            cell.border = StyleConfig.THIN_BORDER
        
        # 计算统计数据
        stats = {
            '炸弹总数': [],
            '万能牌数': [],
            '顺子数量': [],
            '对子数量': [],
            '三条数量': [],
            '贡献分': [],
            '连炸贡献分': [],
            '单炸贡献分': [],
        }
        for line in range(4, 17):
            stats[f'{line}线'] = []
        
        for result in results:
            for i in range(4):
                bombs = result['players'][i]['bombs']
                jokers = result['players'][i]['jokers']
                sequences = result['players'][i].get('sequences', 0)
                pairs = result['players'][i].get('pairs', 0)
                triplets = result['players'][i].get('triplets', 0)
                score = result['players'][i].get('contribution_score', 0)
                
                stats['炸弹总数'].append(len(bombs))
                stats['万能牌数'].append(jokers)
                # 计算线数分布
                line_counts = {}
                for line in range(4, 17):
                    line_counts[line] = 0
                
                # 区分王炸和普通炸弹
                joker_bombs = [b for b in bombs if b['rank'] in ['👑', '🃏']]
                normal_bombs = [b for b in bombs if b['rank'] not in ['👑', '🃏']]
                
                # 王炸线数 = 王数 + 3
                for jb in joker_bombs:
                    line = jb['size'] + 3
                    if 4 <= line <= 16:
                        line_counts[line] += 1
                
                # 普通炸弹检测连炸
                if normal_bombs:
                    CARD_RANKS = ['3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K', 'A', '2']
                    rank_indices = sorted([CARD_RANKS.index(b['rank']) for b in normal_bombs if b['rank'] in CARD_RANKS])
                    groups = []
                    if rank_indices:
                        current = [rank_indices[0]]
                        for j in range(1, len(rank_indices)):
                            if rank_indices[j] == rank_indices[j-1] + 1:
                                current.append(rank_indices[j])
                            else:
                                groups.append(current)
                                current = [rank_indices[j]]
                        groups.append(current)
                    
                    for group in groups:
                        if len(group) >= 3:
                            # 连炸线数 = 最小炸弹张数 + 连环数
                            min_size = min(b['size'] for b in normal_bombs if CARD_RANKS.index(b['rank']) in group)
                            line = min_size + len(group)
                            if 4 <= line <= 16:
                                line_counts[line] += 1
                        else:
                            # 单个炸弹
                            for idx in group:
                                b_size = next(b['size'] for b in normal_bombs if CARD_RANKS.index(b['rank']) == idx)
                                if 4 <= b_size <= 16:
                                    line_counts[b_size] += 1
                
                for line in range(4, 17):
                    stats[f'{line}线'].append(line_counts[line])
                stats['顺子数量'].append(sequences)
                stats['对子数量'].append(pairs)
                stats['三条数量'].append(triplets)
                stats['贡献分'].append(score)
                detail = result['players'][i].get('contribution_detail', {})
                stats['连炸贡献分'].append(detail.get('chain_score', 0))
                stats['单炸贡献分'].append(detail.get('single_score', 0))
        
        # 写入统计数据（跳过空值）
        row += 1
        for stat_name, values in stats.items():
            # 跳过总和为0的统计
            if sum(values) == 0:
                continue
            player_values = [values[i*4:(i+1)*4] for i in range(len(results))]
            
            # 每个玩家的平均值
            for i in range(4):
                player_vals = [pv[i] for pv in player_values]
                ws.cell(row=row, column=i+2).value = round(sum(player_vals) / len(results), 2)
            
            # 总计和平均
            all_values = [v for sublist in player_values for v in sublist]
            ws.cell(row=row, column=6).value = sum(all_values)
            ws.cell(row=row, column=7).value = round(sum(all_values) / (len(results) * 4), 2)
            ws.cell(row=row, column=8).value = max(all_values)
            
            # 指标名称
            ws.cell(row=row, column=1).value = stat_name
            
            # 设置样式
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.font = StyleConfig.NORMAL_FONT
                cell.alignment = StyleConfig.CENTER_ALIGN
                cell.border = StyleConfig.THIN_BORDER
                if row % 2 == 1:
                    cell.fill = StyleConfig.ALT_FILL
            
            row += 1
        
        # === 第二部分：炸弹分布 ===
        row += 2
        ws.merge_cells(f'A{row}:H{row}')
        section_title = ws[f'A{row}']
        section_title.value = '📈 炸弹线数分布'
        section_title.font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')
        section_title.alignment = StyleConfig.LEFT_ALIGN
        
        # 表头
        row += 1
        headers = ['炸弹大小', '出现次数', '占比', '平均每局', '说明', '', '', '']
        ws.append(headers)
        
        # 设置表头样式
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.font = StyleConfig.HEADER_FONT
            cell.fill = StyleConfig.HEADER_FILL
            cell.alignment = StyleConfig.CENTER_ALIGN
            cell.border = StyleConfig.THIN_BORDER
        
        # 统计炸弹线数分布（考虑连炸）
        CARD_RANKS = ['3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K', 'A', '2']
        line_counter = {}
        for line in range(4, 17):
            line_counter[line] = 0
        
        for result in results:
            for player in result['players']:
                bombs = player['bombs']
                joker_bombs = [b for b in bombs if b['rank'] in ['👑', '🃏']]
                normal_bombs = [b for b in bombs if b['rank'] not in ['👑', '🃏']]
                
                # 王炸线数
                for jb in joker_bombs:
                    line = jb['size'] + 3
                    if 4 <= line <= 16:
                        line_counter[line] += 1
                
                # 普通炸弹：检测连炸
                if normal_bombs:
                    rank_indices = sorted([CARD_RANKS.index(b['rank']) for b in normal_bombs if b['rank'] in CARD_RANKS])
                    groups = []
                    if rank_indices:
                        current = [rank_indices[0]]
                        for j in range(1, len(rank_indices)):
                            if rank_indices[j] == rank_indices[j-1] + 1:
                                current.append(rank_indices[j])
                            else:
                                groups.append(current)
                                current = [rank_indices[j]]
                        groups.append(current)
                    
                    for group in groups:
                        if len(group) >= 3:
                            min_size = min(b['size'] for b in normal_bombs if CARD_RANKS.index(b['rank']) in group)
                            line = min_size + len(group)
                            if 4 <= line <= 16:
                                line_counter[line] += 1
                        else:
                            for idx in group:
                                b_size = next(b['size'] for b in normal_bombs if CARD_RANKS.index(b['rank']) == idx)
                                if 4 <= b_size <= 16:
                                    line_counter[b_size] += 1
        
        total_bombs = sum(line_counter.values())
        
        # 说明文字
        desc_map = {
            4: '普通炸弹 (无贡献分)',
            5: '5线 (×2)',
            6: '6线 (×4)',
        }
        for line in range(7, 17):
            desc_map[line] = f'{line}线 (×{2**(line-4)})'
        
        # 写入数据（只显示有值的线数）
        row += 1
        for line in range(4, 17):
            count = line_counter[line]
            if count == 0:
                continue
            percentage = count / total_bombs * 100 if total_bombs > 0 else 0
            per_game = count / len(results)
            desc = desc_map.get(line, '')
            
            ws.cell(row=row, column=1).value = f'{line}线'
            ws.cell(row=row, column=2).value = count
            ws.cell(row=row, column=3).value = f'{percentage:.2f}%'
            ws.cell(row=row, column=4).value = round(per_game, 2)
            ws.cell(row=row, column=5).value = desc
            
            # 设置样式
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.font = StyleConfig.NORMAL_FONT
                cell.alignment = StyleConfig.CENTER_ALIGN
                cell.border = StyleConfig.THIN_BORDER
                if row % 2 == 0:
                    cell.fill = StyleConfig.ALT_FILL
            
            row += 1
        
        # === 第三部分：队伍平衡 ===
        row += 2
        ws.merge_cells(f'A{row}:H{row}')
        section_title = ws[f'A{row}']
        section_title.value = '⚖️ 队伍平衡分析'
        section_title.font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')
        section_title.alignment = StyleConfig.LEFT_ALIGN
        
        # 表头
        row += 1
        headers = ['队伍', '平均炸弹数', '平均万能牌数', '胜率估算', '', '', '', '']
        ws.append(headers)
        
        # 设置表头样式
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = StyleConfig.HEADER_FONT
            cell.fill = StyleConfig.HEADER_FILL
            cell.alignment = StyleConfig.CENTER_ALIGN
            cell.border = StyleConfig.THIN_BORDER
        
        # 统计队伍数据
        team0_bombs = []
        team1_bombs = []
        team0_jokers = []
        team1_jokers = []
        
        for result in results:
            team0_bomb = sum(len(result['players'][i]['bombs']) for i in [0, 2])
            team1_bomb = sum(len(result['players'][i]['bombs']) for i in [1, 3])
            team0_joker = sum(result['players'][i]['jokers'] for i in [0, 2])
            team1_joker = sum(result['players'][i]['jokers'] for i in [1, 3])
            
            team0_bombs.append(team0_bomb)
            team1_bombs.append(team1_bomb)
            team0_jokers.append(team0_joker)
            team1_jokers.append(team1_joker)
        
        # 写入数据
        row += 1
        for team_name, bombs, jokers in [
            ('队伍 0 (玩家 1、3)', team0_bombs, team0_jokers),
            ('队伍 1 (玩家 2、4)', team1_bombs, team1_jokers)
        ]:
            avg_bombs = sum(bombs) / len(results)
            avg_jokers = sum(jokers) / len(results)
            
            # 简单的胜率估算（基于炸弹数）
            total_bombs = sum(team0_bombs) + sum(team1_bombs)
            if total_bombs > 0:
                win_rate = sum(bombs) / total_bombs * 100
            else:
                win_rate = 50
            
            ws.cell(row=row, column=1).value = team_name
            ws.cell(row=row, column=2).value = round(avg_bombs, 2)
            ws.cell(row=row, column=3).value = round(avg_jokers, 2)
            ws.cell(row=row, column=4).value = f'{win_rate:.1f}%'
            
            # 设置样式
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = StyleConfig.NORMAL_FONT
                cell.alignment = StyleConfig.CENTER_ALIGN
                cell.border = StyleConfig.THIN_BORDER
                if row % 2 == 1:
                    cell.fill = StyleConfig.ALT_FILL
            
            row += 1
        
        # === 第四部分：典型牌例 ===
        row += 2
        ws.merge_cells(f'A{row}:H{row}')
        section_title = ws[f'A{row}']
        section_title.value = '🎴 典型牌例（随机 5 局）'
        section_title.font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')
        section_title.alignment = StyleConfig.LEFT_ALIGN
        
        # 随机抽取 5 局
        samples = random.sample(results, min(5, len(results)))
        
        row += 1
        for idx, result in enumerate(samples):
            # 局数标题
            ws.merge_cells(f'A{row}:H{row}')
            game_title = ws[f'A{row}']
            game_title.value = f'第 {idx + 1} 局'
            game_title.font = Font(name='微软雅黑', size=11, bold=True)
            game_title.alignment = StyleConfig.LEFT_ALIGN
            
            # 每个玩家的完整手牌
            for i, player in enumerate(result['players']):
                row += 1
                player_name = f'玩家{i+1} (队伍{player["team"]})'
                bombs_str = ', '.join([f"{b['rank'] * b['size']}" for b in player['bombs']])
                jokers_str = f"万能牌×{player['jokers']}"
                hand_str = player.get('hand_cards_str', '')
                
                ws.cell(row=row, column=1).value = player_name
                ws.cell(row=row, column=2).value = f'炸弹：{len(player["bombs"])}个'
                ws.cell(row=row, column=3).value = jokers_str
                
                row += 1
                ws.merge_cells(f'A{row}:H{row}')
                ws.cell(row=row, column=1).value = f'手牌：{hand_str if hand_str else "无"}'
                row += 1
                ws.merge_cells(f'A{row}:H{row}')
                ws.cell(row=row, column=1).value = f'炸弹详情：{bombs_str if bombs_str else "无"}'
        
        # 调整列宽
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 12
        ws.column_dimensions['A'].width = 15
        
        # 生成文件名
        filename = self.generate_filename()
        filepath = Path(output_dir) / filename
        
        # 保存文件
        self.wb.save(filepath)
        
        return str(filepath)


# ==================== 批量模拟 ====================

class BatchSimulator:
    """批量模拟器"""
    
    def __init__(self, config: Dict):
        self.config = config
        self.num_games = config.get('simulation_config', {}).get('num_games', 100)
        self.min_bombs = config.get('dealing_config', {}).get('min_bombs_per_player', 2)
        self.jokers_per_player = config.get('dealing_config', {}).get('jokers_per_player')
        self.bomb_size_range = config.get('dealing_config', {}).get('bomb_size_range')
        self.seed = config.get('simulation_config', {}).get('random_seed')
        
        if self.seed is not None:
            random.seed(self.seed)
    
    def simulate_one_game(self) -> Dict:
        """模拟一局游戏"""
        dealer = ShuangkouDealer(num_players=4)
        dealer.deal(min_bombs_per_player=self.min_bombs, bomb_size_range=self.bomb_size_range, jokers_per_player=self.jokers_per_player, verbose=False)
        
        result = {
            'players': []
        }
        
        for player in dealer.players:
            bombs = player.hand.find_bombs_with_jokers()
            jokers = sum(1 for card in player.hand.cards if card.rank in ['👑', '🃏'])
            effective_bombs = player.hand.count_effective_bombs()
            # 保存完整手牌列表（中文显示）
            hand_cards_str = ', '.join(str(card) for card in player.hand.cards)
            
            player_data = {
                'id': player.id,
                'team': player.team,
                'bombs': [{'rank': b.rank, 'size': b.size} for b in bombs],
                'effective_bombs': effective_bombs,
                'jokers': jokers,
                'total_cards': player.hand.total_cards(),
                'hand_cards_str': hand_cards_str,  # 完整手牌字符串
                'sequences': player.hand.count_sequences(),
                'pairs': player.hand.count_pairs(),
                'triplets': player.hand.count_triplets(),
                'contribution_score': player.hand.calc_contribution_score(),
                'contribution_detail': player.hand.get_contribution_detail()
            }
            result['players'].append(player_data)
        
        return result
    
    def run(self) -> List[Dict]:
        """运行批量模拟"""
        print(f"🎮 开始模拟 {self.num_games} 局游戏...")
        
        results = []
        for i in range(self.num_games):
            result = self.simulate_one_game()
            results.append(result)
            
            if (i + 1) % 10 == 0:
                print(f"  进度：{i + 1}/{self.num_games} 局")
        
        print(f"✅ 模拟完成！共 {self.num_games} 局")
        return results


# ==================== 主程序 ====================

def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='双扣 - 八王千变 批量发牌器')
    parser.add_argument('--config', type=str, default='config.json', help='配置文件路径')
    parser.add_argument('--output', type=str, default='.', help='输出目录')
    
    args = parser.parse_args()
    
    # 读取配置
    config_path = Path(args.config)
    if not config_path.exists():
        print(f"❌ 配置文件不存在：{config_path}")
        return
    
    print(f"📋 读取配置文件：{config_path}")
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    # 运行模拟
    simulator = BatchSimulator(config)
    results = simulator.run()
    
    # 导出 Excel
    if config.get('output_config', {}).get('export_excel', True):
        print(f"\n📊 生成统计报告...")
        exporter = ExcelExporter(config)
        filepath = exporter.export(results, args.output)
        print(f"✅ Excel 报告已生成：{filepath}")
        print(f"\n💡 文件名规则：双扣统计_{exporter.rule_abbr}_时间戳.xlsx")


if __name__ == '__main__':
    main()
